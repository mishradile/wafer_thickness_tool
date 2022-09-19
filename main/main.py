import openpyxl 
from scipy.interpolate import RBFInterpolator
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from copy import copy
from statistics import stdev
import os
import re
from matplotlib.patches import Circle

wb = openpyxl.load_workbook("./data.xlsx")
#Create image directory to store produced images
if not os.path.isdir("./images/"):
    os.makedirs("./images/") 
else:
    for f in os.listdir("./images/"):
        os.remove(os.path.join("./images/", f))   
errored_sheetnames = []

for ws in wb.worksheets:
    print("Working on "+str(ws))
    # Convert Excel data to pandas DataFrame
    df = pd.DataFrame(ws.values)
    
    #Values set by users
    wafer_radius = float(df.iloc[0,14])
    coord_units = df.iloc[1,14]
    colorbar_min = str(df.iloc[2,14])
    colorbar_max = str(df.iloc[3,14])
    
    #Scale wafer_radius and coord_units appropriately
    scaling_factor =1
    if coord_units=="cm":
        scaling_factor = 0.1
    elif coord_units=="μm":
        scaling_factor = 1000
    
    col_names = df.iloc[0]
    #Drop first row with column names
    df = df.iloc[1:]
    df.dropna(axis=1, how='all', inplace=True)
    df.dropna(axis=0, how='all', inplace=True)
    #Drop the settings columns
    df = df.iloc[:,:-3]
    #Find number of plots to make
    var_num = int(df.shape[1]-2)
    
    #Circle for cropping image, take radius to be 150 (+4 for aesthetics so that data points won't be clipped)
    circle = Circle((0, 0), wafer_radius+4, facecolor='none',
             edgecolor=(0, 0, 0), linewidth=1, alpha=1)

    
    #Loop through each column
    for col_no in range(var_num):
        values = df.iloc[:,col_no]
        values = values.to_numpy()
        points = df.iloc[:,-2:]
        points = points.to_numpy()
        #Check for null/non-numeric values in points coordinates and values, ignore errorneous data points, and generate wafer map with remaining values
        valid_points = []
        valid_values = []
        for i in range(len(values)):
            #TODO: there's probably a better way to do this
            if(isinstance(values[i],float) and isinstance(points[i,0], float) and isinstance(points[i,1], float) and (str(points[i,1]) != 'nan') and (str(points[i,0]) != 'nan') and (str(values[i]) != 'nan')):
                valid_points.append(points[i])
                valid_values.append(values[i])
            else:
                if str(ws) not in errored_sheetnames:
                    errored_sheetnames.append(str(ws))
        points = np.array(valid_points)
        values = np.array(valid_values)
        #Scaling factor to adjust for difference in units between coordinates and wafer radius 
        points /= scaling_factor
        
        fig, ax = plt.subplots()
        
        xgrid = np.mgrid[-wafer_radius-5: wafer_radius+5:200j, -wafer_radius-5: wafer_radius+5:200j]
        xflat = xgrid.reshape(2, -1).T
        #possible: gaussian with epsilon 0.01
        yflat = RBFInterpolator(points, values, kernel='linear')(xflat)
        ygrid = yflat.reshape(200, 200)
        
        #.lower() so that user's caps choice won't matter
        vmin = None if colorbar_min.lower() in ["auto", "none"] else int(colorbar_min)
        vmax = None if colorbar_max.lower() in ["auto", "none"] else int(colorbar_max)
        try:
            im = ax.pcolormesh(*xgrid, ygrid, shading='gouraud', cmap='jet', vmin = vmin, vmax = vmax)
        except ValueError:
            print("Error: Please ensure colobar range entered is decimal without spaces.")
            
        # Use line below for point plot to see true value vs interpolated value
        #p = ax.scatter(*points.T, c=values, s=30, ec='k', cmap='jet')
        p = ax.scatter(*points.T, s=10, ec='k', c='black')
        fig.colorbar(im)
        
        #Adding text to plot
        avg = "{:.2f}".format(np.mean(values, axis=0))
        rng = "{:.2f}".format(max(values)-min(values))
        onesig = "{:.2f}".format(100*stdev(values)/np.mean(values, axis=0))

        avg_string = 'Average: '+str(avg)
        rng_string = '±Range: '+str(rng)
        onesig_string = '1Sig: '+ str(onesig)+"%"
        #Figtext instead of .text so that text position is independent of axes (wafer radius)
        plt.figtext(0.05, 0.03, avg_string, fontsize=10, fontweight = 'bold')
        plt.figtext(0.35, 0.03, rng_string, fontsize=10, fontweight = 'bold')
        plt.figtext(0.65, 0.03, onesig_string, fontsize=10, fontweight = 'bold')
        new_circle = copy(circle)
        ax.add_patch(new_circle)
        im.set_clip_path(new_circle)
        #Display value above each data point
        points_1 = pd.DataFrame(points)
        for index, row in points_1.iterrows():
            ax.text(row.iloc[0], row.iloc[1]+2, str( "{:.1f}".format(values[index])), fontsize=5)
            
        #Add triangular notch at bottom of wafer
        plt.plot([-wafer_radius-2], "^:w")
        
        plt.title('')
        
        fig.tight_layout()
        plt.axis('off')
        image_path = "./images/"+ str(re.search(r'.*?\"(.*)".*' , str(ws)).group(1))+"_"+str(col_names[col_no])
        
        plt.savefig(image_path)
        #plt.show()
        plt.close()


if (len(errored_sheetnames) != 0):
    print("Images generated succesfully. Data points with null/non-numeric values as coordinates/values found and ignored in worksheets: ")
    for sheetname in errored_sheetnames:
        print(sheetname)
    print("View results in /images/ folder.")
else: 
    print("Images generated successfully. View results in /images/ folder.")
        


        
        
        
        